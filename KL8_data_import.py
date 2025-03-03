import json
import math as math
import numpy as np
import pandas as pd
import requests
import openpyxl  # 替换 xlwt
import re
import logging
import os
import time
from tqdm import tqdm
from funcs.requestsdata import requests_data, get_latest_issue_from_system

Lottry_ID = 6  # 快乐8的ID为6; 7乐彩 ID为3；双色球 ID为1，

try:
    df_existing = pd.read_excel('快乐8开奖情况.xlsx')  # 替换 xls 为 xlsx
    last_issue_in_excel = int(df_existing['期号'].max())
except FileNotFoundError:
    last_issue_in_excel = 0

latest_issue_in_system = get_latest_issue_from_system(Lottry_ID)
if latest_issue_in_system is None:
    print("无法获取最新期号，程序终止。")
    exit()

KL8_current_2025_times = latest_issue_in_system - 2025000  # Lukcy 8 open times in 2025
KL8_total_issueCount = 65 + 351 + 351 + 351 + 352 + KL8_current_2025_times
# 快乐8在20201028发行第一期，在当年发行了65期，在2021年总共发行了351期， 2022 = 351 ,2023  - 351 截止20220115共发行了65+351+351+351+21=1139，大家可以在运行本代码时根据实际日期修改本变量。

if last_issue_in_excel == latest_issue_in_system:
    print(f"本地数据已是最新，跳过下载。最新期号: {latest_issue_in_system}")
else:
    wb = openpyxl.Workbook()  # 替换 xlwt 为 openpyxl
    sheet = wb.active  # 使用活动工作表
    sheet.title = '快乐8'  # 设置工作表标题
    # 存储表头文件
    row = ["期号", "开奖日期", "前区号码", "后区号码", "总销售额(元)", "奖池金额(元)",
           "选10中10注数", "选10中10奖金", "选10中9注数", "选10中9奖金", "选10中8注数", "选10中8奖金",
           "选10中7注数", "选10中7奖金", "选10中6注数", "选10中6奖金", "选10中5注数", "选10中5奖金",
           "选10中0注数", "选10中0奖金",
           "选9中9注数", "选9中9奖金", "选9中8注数", "选9中8奖金", "选9中7注数", "选9中7奖金",
           "选9中6注数", "选9中6奖金", "选9中5注数", "选9中5奖金", "选9中4注数", "选9中4奖金",
           "选9中0注数", "选9中0奖金",
           "选8中8注数", "选8中8奖金", "选8中7注数", "选8中7奖金", "选8中6注数", "选8中6奖金",
           "选8中5注数", "选8中5奖金", "选8中4注数", "选8中4奖金", "选8中0注数", "选8中0奖金",
           "选7中7注数", "选7中7奖金", "选7中6注数", "选7中6奖金", "选7中5注数", "选7中5奖金",
           "选7中4注数", "选7中4奖金", "选7中0注数", "选7中0奖金",
           "选6中6注数", "选6中6奖金", "选6中5注数", "选6中5奖金", "选6中4注数", "选6中4奖金",
           "选6中3注数", "选6中3奖金",
           "选5中5注数", "选5中5奖金", "选5中4注数", "选5中4奖金", "选5中3注数", "选5中3奖金",
           "选4中4注数", "选4中4奖金", "选4中3注数", "选4中3奖金", "选4中2注数", "选4中2奖金",
           "选3中3注数", "选3中3奖金", "选3中2注数", "选3中2奖金",
           "选2中2注数", "选2中2奖金",
           "选1中1注数", "选1中1奖金"
           ]

    # 写入表头
    for i, title in enumerate(row):
        sheet.cell(row=1, column=i + 1, value=title)  # 修改写入方式

    i = 2  # 修改i的初始值
    range_max = math.floor(KL8_total_issueCount / 30 + 1) if KL8_total_issueCount % 30 == 0 else math.floor(
        KL8_total_issueCount / 30 + 2)
    # 如果issueCount是30的整数倍则range_max=math.floor(issueCount/30+1)，否则range_max=math.floor(issueCount/30+2)
    for pageNum_i in range(1, range_max):  # 页数必须正好，多了就会返回重复数据，431/30=14.3
        tony_dict = requests_data(pageNum_i, KL8_total_issueCount, Lottry_ID)
        for j in tony_dict:
            if j != '{':
                tony_dict = tony_dict[-(len(tony_dict) - 1):]
            else:
                break
        if tony_dict[len(tony_dict) - 1] == ')':
            tony_dict = tony_dict[:len(tony_dict) - 1]  # 删除最后一个右括号)
        content = json.loads(tony_dict)
        content_data = content['data']

        for item in content_data:
            sheet.cell(row=i, column=1, value=item['issue'])
            sheet.cell(row=i, column=2, value=item['openTime'])
            sheet.cell(row=i, column=3, value=item['frontWinningNum'])
            sheet.cell(row=i, column=4, value=item['backWinningNum'])
            sheet.cell(row=i, column=5, value=item['saleMoney'])
            sheet.cell(row=i, column=6, value=item['prizePoolMoney'])

            # 处理中奖详情
            winner_details = item['winnerDetails']
            winner_data = {}
            for detail in winner_details:
                award_etc = detail['awardEtc']
                base_winner = detail['baseBetWinner']
                winner_data[award_etc] = {
                    'awardNum': base_winner['awardNum'],
                    'awardMoney': base_winner['awardMoney']
                }

            # 写入中奖信息
            sheet.cell(row=i, column=7, value=winner_data.get('x10z10', {}).get('awardNum', ''))
            sheet.cell(row=i, column=8, value=winner_data.get('x10z10', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=9, value=winner_data.get('x10z9', {}).get('awardNum', ''))
            sheet.cell(row=i, column=10, value=winner_data.get('x10z9', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=11, value=winner_data.get('x10z8', {}).get('awardNum', ''))
            sheet.cell(row=i, column=12, value=winner_data.get('x10z8', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=13, value=winner_data.get('x10z7', {}).get('awardNum', ''))
            sheet.cell(row=i, column=14, value=winner_data.get('x10z7', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=15, value=winner_data.get('x10z6', {}).get('awardNum', ''))
            sheet.cell(row=i, column=16, value=winner_data.get('x10z6', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=17, value=winner_data.get('x10z5', {}).get('awardNum', ''))
            sheet.cell(row=i, column=18, value=winner_data.get('x10z5', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=19, value=winner_data.get('x10z0', {}).get('awardNum', ''))
            sheet.cell(row=i, column=20, value=winner_data.get('x10z0', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=21, value=winner_data.get('x9z9', {}).get('awardNum', ''))
            sheet.cell(row=i, column=22, value=winner_data.get('x9z9', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=23, value=winner_data.get('x9z8', {}).get('awardNum', ''))
            sheet.cell(row=i, column=24, value=winner_data.get('x9z8', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=25, value=winner_data.get('x9z7', {}).get('awardNum', ''))
            sheet.cell(row=i, column=26, value=winner_data.get('x9z7', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=27, value=winner_data.get('x9z6', {}).get('awardNum', ''))
            sheet.cell(row=i, column=28, value=winner_data.get('x9z6', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=29, value=winner_data.get('x9z5', {}).get('awardNum', ''))
            sheet.cell(row=i, column=30, value=winner_data.get('x9z5', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=31, value=winner_data.get('x9z4', {}).get('awardNum', ''))
            sheet.cell(row=i, column=32, value=winner_data.get('x9z4', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=33, value=winner_data.get('x9z0', {}).get('awardNum', ''))
            sheet.cell(row=i, column=34, value=winner_data.get('x9z0', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=35, value=winner_data.get('x8z8', {}).get('awardNum', ''))
            sheet.cell(row=i, column=36, value=winner_data.get('x8z8', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=37, value=winner_data.get('x8z7', {}).get('awardNum', ''))
            sheet.cell(row=i, column=38, value=winner_data.get('x8z7', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=39, value=winner_data.get('x8z6', {}).get('awardNum', ''))
            sheet.cell(row=i, column=40, value=winner_data.get('x8z6', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=41, value=winner_data.get('x8z5', {}).get('awardNum', ''))
            sheet.cell(row=i, column=42, value=winner_data.get('x8z5', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=43, value=winner_data.get('x8z4', {}).get('awardNum', ''))
            sheet.cell(row=i, column=44, value=winner_data.get('x8z4', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=45, value=winner_data.get('x8z0', {}).get('awardNum', ''))
            sheet.cell(row=i, column=46, value=winner_data.get('x8z0', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=47, value=winner_data.get('x7z7', {}).get('awardNum', ''))
            sheet.cell(row=i, column=48, value=winner_data.get('x7z7', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=49, value=winner_data.get('x7z6', {}).get('awardNum', ''))
            sheet.cell(row=i, column=50, value=winner_data.get('x7z6', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=51, value=winner_data.get('x7z5', {}).get('awardNum', ''))
            sheet.cell(row=i, column=52, value=winner_data.get('x7z5', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=53, value=winner_data.get('x7z4', {}).get('awardNum', ''))
            sheet.cell(row=i, column=54, value=winner_data.get('x7z4', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=55, value=winner_data.get('x7z0', {}).get('awardNum', ''))
            sheet.cell(row=i, column=56, value=winner_data.get('x7z0', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=57, value=winner_data.get('x6z6', {}).get('awardNum', ''))
            sheet.cell(row=i, column=58, value=winner_data.get('x6z6', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=59, value=winner_data.get('x6z5', {}).get('awardNum', ''))
            sheet.cell(row=i, column=60, value=winner_data.get('x6z5', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=61, value=winner_data.get('x6z4', {}).get('awardNum', ''))
            sheet.cell(row=i, column=62, value=winner_data.get('x6z4', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=63, value=winner_data.get('x6z3', {}).get('awardNum', ''))
            sheet.cell(row=i, column=64, value=winner_data.get('x6z3', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=65, value=winner_data.get('x5z5', {}).get('awardNum', ''))
            sheet.cell(row=i, column=66, value=winner_data.get('x5z4', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=67, value=winner_data.get('x5z3', {}).get('awardNum', ''))
            sheet.cell(row=i, column=68, value=winner_data.get('x5z3', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=69, value=winner_data.get('x4z4', {}).get('awardNum', ''))
            sheet.cell(row=i, column=70, value=winner_data.get('x4z4', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=71, value=winner_data.get('x4z3', {}).get('awardNum', ''))
            sheet.cell(row=i, column=72, value=winner_data.get('x4z3', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=73, value=winner_data.get('x4z2', {}).get('awardNum', ''))
            sheet.cell(row=i, column=74, value=winner_data.get('x4z2', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=75, value=winner_data.get('x3z3', {}).get('awardNum', ''))
            sheet.cell(row=i, column=76, value=winner_data.get('x3z3', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=77, value=winner_data.get('x3z2', {}).get('awardNum', ''))
            sheet.cell(row=i, column=78, value=winner_data.get('x3z2', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=79, value=winner_data.get('x2z2', {}).get('awardNum', ''))
            sheet.cell(row=i, column=80, value=winner_data.get('x2z2', {}).get('awardMoney', ''))
            sheet.cell(row=i, column=81, value=winner_data.get('x1z1', {}).get('awardNum', ''))
            sheet.cell(row=i, column=82, value=winner_data.get('x1z1', {}).get('awardMoney', ''))

            i = i + 1
# 保存
wb.save("快乐8开奖情况.xlsx")  # 修改保存的文件格式
print("数据已成功下载并保存。")

# 统计每年的开奖次数
# 读取 Excel 文件
df = pd.read_excel('快乐8开奖情况.xlsx', header=0)  # 修改读取的文件格式

# 将“开奖日期”列转换为 datetime 类型
df['开奖日期'] = pd.to_datetime(df['开奖日期'])

# 提取年份
df['年份'] = df['开奖日期'].dt.year

# 按年份分组并统计期数
yearly_counts = df.groupby('年份')['期号'].count()

# 打印结果
print(yearly_counts)