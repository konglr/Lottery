import json
import math
import openpyxl
import logging
import pandas as pd
from tqdm import tqdm
from requestsdata import requests_data, get_latest_issue_from_system

Lottry_ID = 1  # 快乐8的ID为6; 7乐彩 ID为3；福彩3D， ID为2，双色球 ID为1，
# 配置日志记录
logging.basicConfig(filename='my_log_file.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

try:
    df_existing = pd.read_excel('双色球开奖情况.xlsx')
    if df_existing.empty:  # 如果Excel文件存在，但是是空的，则也设置last_issue_in_excel = 0
        last_issue_in_excel = 0
    else:
        last_issue_in_excel = int(df_existing['期号'].max())
except FileNotFoundError:
    last_issue_in_excel = 0

latest_issue_in_system = get_latest_issue_from_system(Lottry_ID)
if latest_issue_in_system is None:
    print("无法获取最新期号，程序终止。")
    exit()

current_2026_times = latest_issue_in_system - 2026000
total_issueCount = 3246 + 151 + current_2026_times

# 如果本地文件最后一期与系统最新期号相同，则跳过下载
if last_issue_in_excel == latest_issue_in_system:
    print(f"本地数据已是最新，跳过下载。最新期号: {latest_issue_in_system}")
else:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '双色球'

    row = ["期号", "开奖日期", "WeekDay", "前区号码", "后区号码", "总销售额(元)", "奖池金额(元)",
           "一等奖注数", "一等奖奖金", "二等奖注数", "二等奖奖金", "三等奖注数", "三等奖奖金",
           "四等奖注数", "四等奖奖金", "五等奖注数", "五等奖奖金", "六等奖注数", "六等奖奖金",
           "红球1", "红球2", "红球3", "红球4", "红球5", "红球6", "蓝球"]  # 添加红球和蓝球列
    for i, title in enumerate(row):
        sheet.cell(row=1, column=i + 1, value=title)

    i = 2
    range_max = math.floor(total_issueCount / 100 + 1) if total_issueCount % 100 == 0 else math.floor(total_issueCount / 100 + 2)
    for pageNum_i in tqdm(range(1, range_max), desc="下载进度"):
        tony_dict = requests_data(pageNum_i, total_issueCount, Lottry_ID)
        for j in tony_dict:
            if j != '{':
                tony_dict = tony_dict[-(len(tony_dict) - 1):]
            else:
                break
        if tony_dict[len(tony_dict) - 1] == ')':
            tony_dict = tony_dict[:len(tony_dict) - 1]
        content = json.loads(tony_dict)
        content_data = content['data']
        for item in content_data:
            sheet.cell(row=i, column=1, value=item['issue'])
            sheet.cell(row=i, column=2, value=item['openTime'])
            sheet.cell(row=i, column=3, value=item['week'])
            sheet.cell(row=i, column=4, value=item['frontWinningNum'])
            sheet.cell(row=i, column=5, value=item['backWinningNum'])
            sheet.cell(row=i, column=6, value=item['saleMoney'])
            sheet.cell(row=i, column=7, value=item['prizePoolMoney'])
            winner_details = item.get('winnerDetails', [])
            for award in winner_details:
                award_etc = award.get('awardEtc', '')
                base_bet_winner = award.get('baseBetWinner', {})
                try:
                    award_level = int(award_etc)
                    if 1 <= award_level <= 6:
                        col_index = 8 + (award_level - 1) * 2
                        sheet.cell(row=i, column=col_index, value=base_bet_winner.get('awardNum', ''))
                        sheet.cell(row=i, column=col_index + 1, value=base_bet_winner.get('awardMoney', ''))
                except ValueError:
                    print(f"awardEtc: {award_etc} 不是有效的数字")
                    continue
            # 写入红球和蓝球
            front_nums = item['frontWinningNum'].split()
            back_nums = item['backWinningNum'].split()
            for j in range(6):
                sheet.cell(row=i, column=20 + j, value=int(front_nums[j])) # 修改红球的写入列
            sheet.cell(row=i, column=26, value=int(back_nums[0])) #修改蓝球的写入列
            i += 1
    wb.save("双色球开奖情况.xlsx")
    print("数据已成功下载并保存。")

# 数据检验部分代码
# 读取 Excel 文件
df = pd.read_excel('双色球开奖情况.xlsx', header=0) #替换xls为xlsx

# 检查最早的数据是否为 2003001
earliest_issue = df['期号'].min()
if earliest_issue != 2003001:
    print(f"1.最早的数据是2003001，符合预期。")

# 检查最新一期的数据是否与系统里的相同
# Configure logging (as before)
logging.basicConfig(filename='my_log_file.log', level=logging.INFO)

last_issue_in_excel = df['期号'].max()  # Calculate AFTER reading and processing the Excel file
last_issue_in_excel = int(last_issue_in_excel)  # Convert to int for comparison

# --- Comparison ---
latest_issue_in_system = get_latest_issue_from_system(Lottry_ID)

if latest_issue_in_system is None:
    print("Failed to get latest issue from system.")
else:
    if last_issue_in_excel == int(latest_issue_in_system):  # Convert to int for comparison
        print(f"2.Excel与系统数据同步. 最新期数为: {last_issue_in_excel}")
        logging.info(f"Data synchronized: {last_issue_in_excel}")
    else:
        print(f"WARNING: Excel and system data are NOT synchronized!")
        print(f"Excel: {last_issue_in_excel}, System: {latest_issue_in_system}")
        logging.warning(f"Data mismatch: Excel: {last_issue_in_excel}, System: {latest_issue_in_system}")

# 检查是否有重复数据
duplicate_issues = df[df.duplicated(['期号'], keep=False)]
if not duplicate_issues.empty:
    print("3.警告：发现重复的期号:")
    print(duplicate_issues[['期号']])
else:
    print("3.未发现重复数据。")

# 统计每年的开奖次数
# 读取 Excel 文件
df = pd.read_excel('双色球开奖情况.xlsx', header=0) #替换xls为xlsx

# 将“开奖日期”列转换为 datetime 类型
df['开奖日期'] = pd.to_datetime(df['开奖日期'])

# 提取年份
df['年份'] = df['开奖日期'].dt.year

# 按年份分组并统计期数
yearly_counts = df.groupby('年份')['期号'].count()

# 打印结果
print(yearly_counts)


# 读取历史数据
df = pd.read_excel("双色球开奖情况.xlsx")


# 定义红球列名（假设为6个红球列名）
red_ball_columns = ['红球1', '红球2', '红球3', '红球4', '红球5', '红球6']
# 确保红球列为整数类型
df[red_ball_columns] = df[red_ball_columns].astype(int)

# 计算各项特征
奇数 = []
偶数 = []
小号 = []
大号 = []
一区 = []
二区 = []
三区 = []
重号 = []
邻号 = []
孤号 = []
二连 = []
三连 = []
四连 = []
五连 = []
六连 = []
二跳 = []
三跳 = []
四跳 = []
五跳 = []
六跳 = []
和值 = []
AC = []
跨度 = []

# 遍历每一期数据（倒序遍历，从最新的开始）
for i in range(0, len(df), 1):
    nums = sorted(df.loc[i, red_ball_columns].tolist())  # 当前期红球数据（已排序）

    # 1. 计算奇数和偶数数量
    odd_count = sum(1 for num in nums if num % 2 == 1)
    even_count = 6 - odd_count
    奇数.append(odd_count)
    偶数.append(even_count)

    # 2. 计算小号和大号数量
    small_count = sum(1 for num in nums if num <= 16)
    big_count = 6 - small_count
    小号.append(small_count)
    大号.append(big_count)

    # 3. 计算三区分布
    zone1_count = sum(1 for num in nums if 1 <= num <= 11)
    zone2_count = sum(1 for num in nums if 12 <= num <= 24)
    zone3_count = sum(1 for num in nums if 25 <= num <= 33)
    一区.append(zone1_count)
    二区.append(zone2_count)
    三区.append(zone3_count)

    # 4. 计算重号（与上一期相同的号码）
    if i < len(df) - 1:  # 判断是否有上一期数据
        last_nums = sorted(df.loc[i + 1, red_ball_columns].tolist())  # 上一期红球
    else:
        last_nums = []  # 如果没有上一期数据，则设为空
    repeat_count = len(set(nums) & set(last_nums))
    重号.append(repeat_count)

    # 5. 计算邻号（与上一期号码相邻的号码）
    adjacent_count = sum(1 for num in nums if((num - 1 in last_nums) or (num + 1 in last_nums)))
    邻号.append(adjacent_count)

    # 6. 计算孤号（去掉重号和邻号后剩下的号码）
    孤号.append(6 - repeat_count - adjacent_count)

    # 7. 计算连号
    def count_consecutive(nums):
        """统计连号，解决重复计算问题"""
        nums.sort()  # 确保数字排序
        n = len(nums)
        counts = {'二连': 0, '三连': 0, '四连': 0, '五连': 0, '六连': 0}
        i = 0
        while i < n - 1:
            if nums[i] + 1 == nums[i + 1]:
                length = 2
                while i + length < n and nums[i + length - 1] + 1 == nums[i + length]:
                    length += 1

                if length == 2:
                    counts['二连'] += 1
                    i += 2
                elif length == 3:
                    counts['三连'] += 1
                    i += 3
                elif length == 4:
                    counts['四连'] += 1
                    i += 4
                elif length == 5:
                    counts['五连'] += 1
                    i += 5
                elif length == 6:
                    counts['六连'] += 1
                    i += 6
            else:
                i += 1
        return counts

    consecutive_counts = count_consecutive(nums)

    二连.append(consecutive_counts['二连'])
    三连.append(consecutive_counts['三连'])
    四连.append(consecutive_counts['四连'])
    五连.append(consecutive_counts['五连'])
    六连.append(consecutive_counts['六连'])

    # 8. 计算跳号
    def count_jumps(nums):
        """统计跳号，解决重复计算问题"""
        nums.sort()  # 确保数字排序
        n = len(nums)
        counts = {'二跳': 0, '三跳': 0, '四跳': 0, '五跳': 0, '六跳': 0}
        i = 0
        while i < n - 1:
            diff = nums[i + 1] - nums[i]
            if diff >= 2:
                length = 1
                while i + length < n - 1 and nums[i + length + 1] - nums[i + length] == diff:
                    length += 1

                if diff == 2:
                    if length == 1:
                        counts['二跳'] += 1
                        i += 2
                    elif length == 2:
                        counts['三跳'] += 1
                        i += 3
                    elif length == 3:
                        counts['四跳'] += 1
                        i += 4
                    elif length == 4:
                        counts['五跳'] += 1
                        i += 5
                    elif length == 5:
                        counts['六跳'] += 1
                        i += 6
                    else:
                        i += 1
                elif diff == 3:
                    if length == 2:
                        counts['三跳'] += 1
                        i += 3
                    elif length == 3:
                        counts['四跳'] += 1
                        i += 4
                    elif length == 4:
                        counts['五跳'] += 1
                        i += 5
                    elif length == 5:
                        counts['六跳'] += 1
                        i += 6
                    else:
                        i += 1
                elif diff == 4:
                    if length == 3:
                        counts['四跳'] += 1
                        i += 4
                    elif length == 4:
                        counts['五跳'] += 1
                        i += 5
                    elif length == 5:
                        counts['六跳'] += 1
                        i += 6
                    else:
                        i += 1
                elif diff == 5:
                    if length == 4:
                        counts['五跳'] += 1
                        i += 5
                    elif length == 5:
                        counts['六跳'] += 1
                        i += 6
                    else:
                        i += 1
                elif diff == 6:
                    if length == 5:
                        counts['六跳'] += 1
                        i += 6
                    else:
                        i += 1

                else:
                    i += 1
            else:
                i += 1

        return counts

    jump_counts = count_jumps(nums)

    二跳.append(jump_counts['二跳'])
    三跳.append(jump_counts['三跳'])
    四跳.append(jump_counts['四跳'])
    五跳.append(jump_counts['五跳'])
    六跳.append(jump_counts['六跳'])

    # 9. 计算和值
    sum_value = sum(nums)
    和值.append(sum_value)

    # 10. 计算 AC 值
    differences = sorted(set(abs(a - b) for a in nums for b in nums if a > b))
    ac_value = len(differences) - 5
    AC.append(ac_value)

    # 11. 计算跨度
    span_value = max(nums) - min(nums)
    跨度.append(span_value)

# 将计算结果加入 DataFrame
df["奇数"] = 奇数
df["偶数"] = 偶数
df["小号"] = 小号
df["大号"] = 大号
df["一区"] = 一区
df["二区"] = 二区
df["三区"] = 三区
df["重号"] = 重号
df["邻号"] = 邻号
df["孤号"] = 孤号
df["二连"] = 二连
df["三连"] = 三连
df["四连"] = 四连
df["五连"] = 五连
df["六连"] = 六连
df["二跳"] = 二跳
df["三跳"] = 三跳
df["四跳"] = 四跳
df["五跳"] = 五跳
df["六跳"] = 六跳
df["和值"] = 和值
df["AC"] = AC
df["跨度"] = 跨度

# 保存更新后的数据到 Excel
df.to_excel("双色球开奖情况.xlsx", index=False)

print("数据处理完成，并已写入 Excel！")

df.head(10)